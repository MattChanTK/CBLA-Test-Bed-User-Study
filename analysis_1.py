__author__ = 'Matthew'
# This analysis performs analysis on the survey data

import openpyxl as pyxl
import os
import numpy as np
from scipy import stats

# define the root directory where the data are
data_root_dir = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
print("Study Data Root Directory: ", data_root_dir, end='\n\n')

# remember where the program executed
program_dir = os.getcwd()

# change to the study data directory
os.chdir(data_root_dir)

# open the workbook containing the data
study_book_file_name = "user_study_dec_7_2015.xlsx"
study_book = pyxl.load_workbook(study_book_file_name)

# open the survey sheet
interest_sheet = study_book.get_sheet_by_name('Interest Level')

# define the table's indices
CBLA_FIRST_COL = 2
DATA_FIRST_COL = 3
DATA_MID_COL = 6
DATA_LAST_COL = 10
DATA_FIRST_ROW = 2
DATA_LAST_ROW = interest_sheet.get_highest_row()

# Put Data into four groups
prescripted_first = []
prescripted_second = []
cbla_first = []
cbla_second = []

for i in range(DATA_FIRST_ROW, DATA_LAST_ROW+1):

    # getting the first half of the data
    sample_data_1 = []
    for j in range(DATA_FIRST_COL, DATA_MID_COL+1):
        sample_data_1.append(interest_sheet.cell(row=i, column=j).value)

    # getting the second half of the data
    sample_data_2 = []
    for j in range(DATA_MID_COL+1, DATA_LAST_COL+1):
        sample_data_2.append(interest_sheet.cell(row=i, column=j).value)

    # if this is a CBLA first kind of trial
    if interest_sheet.cell(row=i, column=CBLA_FIRST_COL).value:
        cbla_first.append(tuple(sample_data_1))
        prescripted_second.append(tuple(sample_data_2))

    # if it is a prescripted first kind of trial
    else:
        prescripted_first.append(tuple(sample_data_1))
        cbla_second.append(tuple(sample_data_2))

# display the different groups of data
# print("Prescripted Mode (first half):\t", prescripted_first)
# print("Prescripted Mode (second half):\t", prescripted_second)
# print("CBLA Mode (first half):\t\t\t", cbla_first)
# print("CBLA Mode (second half):\t\t", cbla_second)

# store them in a dictionary
prescripted_first_avg = [np.mean(x) for x in prescripted_first]
prescripted_second_avg = [np.mean(x) for x in prescripted_second]
cbla_first_avg = [np.mean(x) for x in cbla_first]
cbla_second_avg = [np.mean(x) for x in cbla_second]


# display the different groups of averaged data
print("Average Interest Level (from Questionaire Cards)")
print("=====================================", end='\n\n')
print("Prescripted Mode (first half):\t", prescripted_first_avg)
print("Prescripted Mode (second half):\t", prescripted_second_avg)
print("CBLA Mode (first half):\t\t\t", cbla_first_avg)
print("CBLA Mode (second half):\t\t", cbla_second_avg)
print("")

# T-Test
p_significant = 0.10

print("Welch's T-Test for unequal variance")
print("=====================================", end='\n\n')

# prescripted: test if it matters if it was in the first-half or second-half
print("1. Prescripted Mode: on during first half vs one during second half")
# Null Hypothesis
print("--- H0: Average interest level for prescripted mode is the same regardless if it is on first or second.")
# Alternate Hypothesis
print("--- H1: Average interest level for prescripted mode is different depending on whether it is on first or second.")
# Test Data
sample_1 = prescripted_first_avg
sample_2 = prescripted_second_avg
print("Sample 1:  ", sample_1, " mean = ", np.mean(sample_1))
print("Sample 2:  ", sample_2, " mean = ", np.mean(sample_2))
# Analysis
t_stat, p_val = stats.ttest_ind(sample_1, sample_2, equal_var=False)
print("T-Stat: ", t_stat ,"; P-Value: ", p_val, end='  --->')
if p_val > p_significant:
    print("H0 cannot be rejected.\n")
else:
    print("H0 rejected.")

    if t_stat > 0:
        compare_word = 'more'
    else:
        compare_word = 'less'
    print("For prescripted Mode, the first segment is %s interesting than second segment with %.2f%% confidence.\n"
          % (compare_word, 100-p_val/2*100))
# CBLA: test if it matters if it was in the first-half or second-half
print("2. CBLA Mode: on during first half vs one during second half")
# Null Hypothesis
print("--- H0: Average interest level for CBLA mode is the same regardless if it is on first or second.")
# Alternate Hypothesis
print("--- H1: Average interest level for CBLA mode is different depending on whether it is on first or second.")
# Test Data
sample_1 = cbla_first_avg
sample_2 = cbla_second_avg
print("Sample 1:  ", sample_1, " mean = ", np.mean(sample_1))
print("Sample 2:  ", sample_2, " mean = ", np.mean(sample_2))
# Analysis
t_stat, p_val = stats.ttest_ind(sample_1, sample_2, equal_var=False)
print("T-Stat: ", t_stat ,"; P-Value: ", p_val, end='  --->')
if p_val > p_significant:
    print("H0 cannot be rejected.\n")
else:
    print("H0 rejected.")

    if t_stat > 0:
        compare_word = 'more'
    else:
        compare_word = 'less'
    print("For CBLA Mode, the first segment is %s interesting than second segment with %.2f%% confidence.\n"
          % (compare_word, 100-p_val/2*100))


# CBLA: test if it matters if it was in the first-half or second-half
print("3. Both Modes: on during first half vs one during second half")
# Null Hypothesis
print("--- H0: Average interest level is the same regardless if it is on first or second.")
# Alternate Hypothesis
print("--- H1: Average interest level is different depending on whether it is on first or second.")
# Test Data
sample_1 = cbla_first_avg+prescripted_first_avg
sample_2 = cbla_second_avg+prescripted_second_avg
print("Sample 1:  ", sample_1, " mean = ", np.mean(sample_1))
print("Sample 2:  ", sample_2, " mean = ", np.mean(sample_2))
# Analysis
t_stat, p_val = stats.ttest_ind(sample_1, sample_2, equal_var=False)
print("T-Stat: ", t_stat ,"; P-Value: ", p_val, end='  --->')
if p_val > p_significant:
    print("H0 cannot be rejected.\n")
else:
    print("H0 rejected.")

    if t_stat > 0:
        compare_word = 'more'
    else:
        compare_word = 'less'
    print("The first segment is %s interesting than second segment with %.2f%% confidence.\n"
          % (compare_word, 100-p_val/2*100))

# prescripted vs cbla: for first half
print("4. CBLA Mode vs. Prescripted Mode: on during first half")
# Null Hypothesis
print("--- H0: Average interest level for CBLA mode is the same as Prescripted mode when it is on first.")
# Alternate Hypothesis
print("--- H1: Average interest level for CBLA mode is different from Prescripted mode when it is on first.")
# Test Data
sample_1 = prescripted_first_avg
sample_2 = cbla_first_avg
print("Sample 1:  ", sample_1, " mean = ", np.mean(sample_1))
print("Sample 2:  ", sample_2, " mean = ", np.mean(sample_2))
# Analysis
t_stat, p_val = stats.ttest_ind(sample_1, sample_2, equal_var=False)
print("T-Stat: ", t_stat ,"; P-Value: ", p_val, end='  --->')
if p_val > p_significant:
    print("H0 cannot be rejected.\n")
else:
    print("H0 rejected.")

    if t_stat > 0:
        compare_word = 'more'
    else:
        compare_word = 'less'
    print("Prescripted Mode is %s interesting than CBLA mode when it is on during the first segment with %.2f%% confidence.\n"
          % (compare_word, 100-p_val/2*100))

# prescripted vs cbla: for second half
print("5. CBLA Mode vs. Prescripted Mode: on during second half")
# Null Hypothesis
print("--- H0: Average interest level for CBLA mode is the same as Prescripted mode when it is on second.")
# Alternate Hypothesis
print("--- H1: Average interest level for CBLA mode is different from Prescripted mode when it is on second.")
# Test Data
sample_1 = prescripted_second_avg
sample_2 = cbla_second_avg
print("Sample 1:  ", sample_1, " mean = ", np.mean(sample_1))
print("Sample 2:  ", sample_2, " mean = ", np.mean(sample_2))
# Analysis
t_stat, p_val = stats.ttest_ind(sample_1, sample_2, equal_var=False)
print("T-Stat: ", t_stat ,"; P-Value: ", p_val, end='  --->')
if p_val > p_significant:
    print("H0 cannot be rejected.\n")
else:
    print("H0 rejected.")

    if t_stat > 0:
        compare_word = 'more'
    else:
        compare_word = 'less'
    print("Prescripted Mode is %s interesting than CBLA mode when it is on during the second segment with %.2f%% confidence.\n"
          % (compare_word, 100-p_val/2*100))

# prescripted vs cbla: for both halves
print("6. CBLA Mode vs. Prescripted Mode: on during the whole trial")
# Null Hypothesis
print("--- H0: Average interest level for CBLA mode is the same as Prescripted mode.")
# Alternate Hypothesis
print("--- H1: Average interest level for CBLA mode is different from Prescripted mode.")
# Test Data
sample_1 = prescripted_first_avg + prescripted_second_avg
sample_2 = cbla_first_avg + cbla_second_avg
print("Sample 1:  ", sample_1, " mean = ", np.mean(sample_1))
print("Sample 2:  ", sample_2, " mean = ", np.mean(sample_2))
# Analysis
t_stat, p_val = stats.ttest_ind(sample_1, sample_2, equal_var=False)
print("T-Stat: ", t_stat ,"; P-Value: ", p_val, end='  --->')
if p_val > p_significant:
    print("H0 cannot be rejected.\n")
else:
    print("H0 rejected.")

    if t_stat > 0:
        compare_word = 'more'
    else:
        compare_word = 'less'
    print("Prescripted Mode is %s interesting than CBLA mode overall with %.2f%% confidence.\n"
          % (compare_word, 100-p_val/2*100))



# T-Test Related Sample
print("\nPaired T-Test")
print("=====================================", end='\n\n')

# prescripted vs cbla: for prescripted first
print("1. CBLA Mode vs. Prescripted Mode when prescripted is on first")
# Null Hypothesis
print("--- H0: Participants find that CBLA Mode is equally as interesting as prescripted mode when prescripted Mode is on first")
# Alternate Hypothesis
print("--- H1: Participants find that CBLA Mode is not equally as interesting as prescripted mode when prescripted Mode is on first")
# Test Data
sample_1 = cbla_second_avg
sample_2 = prescripted_first_avg
print("Sample 1:  ", sample_1, " mean = ", np.mean(sample_1))
print("Sample 2:  ", sample_2, " mean = ", np.mean(sample_2))
# Analysis
t_stat, p_val = stats.ttest_rel(sample_1, sample_2)
print("T-Stat: ", t_stat ,"; P-Value: ", p_val, end='  --->')
if p_val > p_significant:
    print("H0 cannot be rejected.\n")
else:
    print("H0 rejected.")

    if t_stat > 0:
        compare_word = 'more'
    else:
        compare_word = 'less'
    print("CBLA Mode is %s interesting than Prescripted mode when Prescripted Mode is on first with %.2f%% confidence.\n"
          % (compare_word, 100-p_val/2*100))


# prescripted vs cbla: for CBLA first
print("2. CBLA Mode vs. Prescripted Mode when CBLA is on first")
# Null Hypothesis
print("--- H0: Participants find that CBLA Mode is equally as interesting as prescripted mode when CBLA Mode is on first")
# Alternate Hypothesis
print("--- H1: Participants find that CBLA Mode is not equally as interesting as prescripted mode when CBLA Mode is on first")
# Test Data
sample_1 = cbla_first_avg
sample_2 = prescripted_second_avg
print("Sample 1:  ", sample_1, " mean = ", np.mean(sample_1))
print("Sample 2:  ", sample_2, " mean = ", np.mean(sample_2))
# Analysis
t_stat, p_val = stats.ttest_rel(sample_1, sample_2)
print("T-Stat: ", t_stat ,"; P-Value: ", p_val, end='  --->')
if p_val > p_significant:
    print("H0 cannot be rejected.\n")
else:
    print("H0 rejected.")

    if t_stat > 0:
        compare_word = 'more'
    else:
        compare_word = 'less'
    print("CBLA Mode is %s interesting than prescripted mode when CBLA Mode is on first with %.2f%% confidence.\n"
          % (compare_word, 100-p_val/2*100))


# First segment vs second segment for both modes
print("3. First Segment vs. Second Segment (for both modes)")
# Null Hypothesis
print("--- H0: Participants find that the first segment is equally as interesting as second segment")
# Alternate Hypothesis
print("--- H1: Participants find that the first segment is not equally as interesting as the second segment")
# Test Data
sample_1 = cbla_first_avg+prescripted_first_avg
sample_2 = prescripted_second_avg+cbla_second_avg
print("Sample 1:  ", sample_1, " mean = ", np.mean(sample_1))
print("Sample 2:  ", sample_2, " mean = ", np.mean(sample_2))
# Analysis
t_stat, p_val = stats.ttest_rel(sample_1, sample_2)
print("T-Stat: ", t_stat ,"; P-Value: ", p_val, end='  --->')
if p_val > p_significant:
    print("H0 cannot be rejected.\n")
else:
    print("H0 rejected.")

    if t_stat > 0:
        compare_word = 'more'
    else:
        compare_word = 'less'
    print("The first segment is %s interesting than the second segment with %.2f%% confidence.\n"
          % (compare_word, 100-p_val/2*100))

# CBLA Mode vs Prescripted Mode
print("4. CBLA Mode vs Prescripted Mode (for both segments)")
# Null Hypothesis
print("--- H0: Participants find that CBLA Mode is equally as interesting as Prescripted Mode")
# Alternate Hypothesis
print("--- H1: Participants find that CBLA Mode is not equally as interesting as the Prescripted Mode")
# Test Data
sample_1 = cbla_first_avg + cbla_second_avg
sample_2 = prescripted_second_avg +prescripted_first_avg
print("Sample 1:  ", sample_1, " mean = ", np.mean(sample_1))
print("Sample 2:  ", sample_2, " mean = ", np.mean(sample_2))
# Analysis
t_stat, p_val = stats.ttest_rel(sample_1, sample_2)
print("T-Stat: ", t_stat ,"; P-Value: ", p_val, end='  --->')
if p_val > p_significant:
    print("H0 cannot be rejected.\n")
else:
    print("H0 rejected.")

    if t_stat > 0:
        compare_word = 'more'
    else:
        compare_word = 'less'
    print("The CBLA Mode is %s interesting than the Prescripted Mode with %.2f%% confidence.\n"
          % (compare_word, 100-p_val/2*100))