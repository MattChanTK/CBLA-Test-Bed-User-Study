__author__ = 'Matthew'

import xlrd
import openpyxl as pyxl
import csv
import os
import glob
import datetime
import matplotlib.pyplot as plt
import numpy as np
from scipy import stats

# define the root directory where the data are
data_root_dir = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
print("Study Data Root Directory: ", data_root_dir, end='\n\n')


# remember where the program executed
program_dir = os.getcwd()

# change to the study data directory
os.chdir(data_root_dir)

# study's data folder name header
trial_data_dir = "user_study_excel_data"

# change to the study's data folder
os.chdir(trial_data_dir)

# open the workbooks containing the data
study_id = 1
win_size = 10.0

# first, open the activation workbook
activation_book_file_name = "study_%d (%.2fs).xls" % (study_id, win_size)
activation_book = xlrd.open_workbook(activation_book_file_name)

# open the survey sheet
activation_sheet_raw = activation_book.sheet_by_name('Node Activation')

# convert cell to array
activation_sheet = [None]*activation_sheet_raw.nrows
for row_id in range(activation_sheet_raw.nrows):
    row_data = [None]*activation_sheet_raw.ncols
    for col_id in range(activation_sheet_raw.ncols):
        row_data[col_id] = activation_sheet_raw.cell(row_id, col_id).value
    activation_sheet[row_id] = row_data

# second, read in the csv containing the snapshot
os.chdir(data_root_dir)
os.chdir("study_%d" % study_id)
snapshot_sheet = []
with open(glob.glob("cbla*.csv")[0]) as file:
    snapshot_reader = csv.reader(file)
    for row in snapshot_reader:
        snapshot_sheet.append(row)

# third, read in the xlsx containing the survey data
os.chdir(data_root_dir)

# open the workbook containing the data
study_book_file_name = "user_study_dec_7_2015.xlsx"
study_book = pyxl.load_workbook(study_book_file_name)

# open the survey sheet
interest_sheet_raw = study_book.get_sheet_by_name('Interest Level')

# convert cell to array
interest_sheet = [None]*interest_sheet_raw.get_highest_row()
for row_id in range(interest_sheet_raw.get_highest_row()):
    row_data = [None]*interest_sheet_raw.get_highest_column()
    for col_id in range(interest_sheet_raw.get_highest_column()):
        row_data[col_id] = interest_sheet_raw.cell(row=row_id+1, column=col_id+1).value
    interest_sheet[row_id] = row_data

# Specifying indices for the Snapshot sheet
SNAPSHOT_DATA_ROW = 1
SNAPSHOT_CURR_TIME_COL = 1
TRIAL_START_TIME_ROW = 0
TRIAL_START_TIME_COL = 0
ACTIVATION_DATA_ROW = 2
ACTIVATION_DATA_COL = 1
ACTIVATION_TIME_COL = 0
INTEREST_DATA_ROW = 1
INTEREST_DATA_COL = 2

# 1. Acquire the activation values close to the sample points

# acquire the start time of the trial
trial_start_time = activation_sheet[TRIAL_START_TIME_ROW][TRIAL_START_TIME_COL]
trial_start_time = datetime.datetime.strptime(trial_start_time, "%Y-%m-%d %H:%M:%S.%f")

# calculate the time since the start in second
snapshot_win_time = []
for row_id in range(SNAPSHOT_DATA_ROW, len(snapshot_sheet)):
    snapshot_Time = datetime.datetime.strptime(snapshot_sheet[row_id][SNAPSHOT_CURR_TIME_COL], "%Y-%m-%d %H:%M:%S:%f")
    win_time = snapshot_Time - trial_start_time
    snapshot_win_time.append(win_time.total_seconds())

# search for the windows in the activation sheet that are close to the snapshot
sample_activation_all = []
snapshot_id = 0
for row_id in range(ACTIVATION_DATA_ROW, len(activation_sheet)):

    if activation_sheet[row_id][ACTIVATION_TIME_COL] + win_size >= snapshot_win_time[snapshot_id]:
        # cell array before the sample
        cell_array_0 = activation_sheet[row_id][ACTIVATION_DATA_COL:]
        # cell array containing the sample
        try:
            cell_array = activation_sheet[row_id+1][ACTIVATION_DATA_COL:]
        except IndexError:
            cell_array = cell_array_0

        for id, cell in enumerate(cell_array):
            # replace non-numerical values with 0
            if not isinstance(cell, (float, int)):
                cell_array[id] = 0.0
            if not isinstance(cell_array_0[id], (float, int)):
                cell_array_0[id] = 0.0

            # taking average of the value before and during the sample
            cell_array[id] = np.mean((cell_array[id], cell_array_0[id]))

        sample_activation_all.append(cell_array)
        snapshot_id += 1

    if snapshot_id >= len(snapshot_win_time):
        break

# 2. Calculate the Total average activation for each sample point
avg_sample_activation_all = []
for sample_active_level in sample_activation_all:
    avg_sample_activation_all.append(np.mean(sample_active_level))

# 3. Computer the interest level for each sample point (scaled)
sample_interest_level = interest_sheet[study_id][INTEREST_DATA_COL:]
sample_interest_level = [level/45 for level in sample_interest_level]

# 4. Compute average activation for all data points in activation sheet
activation_win_time = []
avg_activation_all = []
for row_id in range(ACTIVATION_DATA_ROW, len(activation_sheet)):
    activation_win_time.append(activation_sheet[row_id][ACTIVATION_TIME_COL])
    cell_array = activation_sheet[row_id][ACTIVATION_DATA_COL:]
    # replace non-numerical values with 0
    for id, cell in enumerate(cell_array):
        if not isinstance(cell, (float, int)):
            cell_array[id] = 0.0
    avg_activation_all.append(np.mean(tuple(cell_array)))


# 6. plot them
print(snapshot_win_time)
print(avg_sample_activation_all)
plt.plot(snapshot_win_time, avg_sample_activation_all, "-ro", ms=10, label="Sample Avg Activation (for all Nodes)")
plt.plot(activation_win_time, avg_activation_all, "-b", label="Avg Activation (for all Nodes)")
plt.plot(snapshot_win_time, sample_interest_level, "-go", ms=10, label="Sample Interest Level")
plt.xlabel("Time (s)")
plt.legend(bbox_to_anchor=(1.05, 1), loc=5, )
plt.show()

# return back to where the program was executed
os.chdir(program_dir)