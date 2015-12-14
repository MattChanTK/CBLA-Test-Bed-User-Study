__author__ = 'Matthew'

import xlrd
import openpyxl as pyxl
import csv
import os
import glob
import datetime
import matplotlib.pyplot as plt
import matplotlib.font_manager as plt_font
import numpy as np
from scipy import stats
from collections import OrderedDict, defaultdict
import save_figure
from test_bed_spatial_map import NodeSpatialMap

# define the root directory where the data are
data_root_dir = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
print("Study Data Root Directory: ", data_root_dir, end='\n\n')


# remember where the program executed
program_dir = os.getcwd()

# window size file to be analyzed
win_size = 1.0

# number of window to be average over for sample activation calculation
sample_win_num = 30
print("Analysis Window = %2.fs" % (win_size*sample_win_num))

# instantiate the map for the CBLA Test Bed
node_map = NodeSpatialMap()
node_map.build_cbla_test_bed_node_map()
node_map.build_cbla_test_bed_grid_map()

# instantiate the dict of data for all studies
correlation_stats = defaultdict(list)

# open the workbooks containing the data
for study_id in range(1, 11):

    # change to the study data directory
    os.chdir(data_root_dir)

    # study's data folder name header
    trial_data_dir = "user_study_excel_data"

    # change to the study's data folder
    os.chdir(trial_data_dir)


    # first, open the activation workbook
    activation_book_file_name = "study_%d (%.2fs).xls" % (study_id, win_size)
    activation_book = xlrd.open_workbook(activation_book_file_name)
    print("Data filename: %s\n" % activation_book_file_name)

    # open the survey sheet
    activation_sheet_raw = activation_book.sheet_by_name('Node Activation')

    # convert cell to array
    activation_sheet = [None]*activation_sheet_raw.nrows
    for row_id in range(activation_sheet_raw.nrows):
        row_data = [None]*activation_sheet_raw.ncols
        for col_id in range(activation_sheet_raw.ncols):
            row_data[col_id] = activation_sheet_raw.cell(row_id, col_id).value
        activation_sheet[row_id] = row_data

    # second, read in the csv containing the snapshot
    os.chdir(data_root_dir)
    os.chdir("study_%d" % study_id)
    snapshot_sheet = []
    with open(glob.glob("cbla*.csv")[0]) as file:
        snapshot_reader = csv.reader(file)
        for row in snapshot_reader:
            snapshot_sheet.append(row)

    # third, read in the xlsx containing the survey data
    os.chdir(data_root_dir)

    # open the workbook containing the data
    study_book_file_name = "user_study_dec_7_2015.xlsx"
    study_book = pyxl.load_workbook(study_book_file_name)

    # open the interest level sheet
    interest_sheet_raw = study_book.get_sheet_by_name('Interest Level')

    # convert cell to array
    interest_sheet = [None]*interest_sheet_raw.get_highest_row()
    for row_id in range(interest_sheet_raw.get_highest_row()):
        row_data = [None]*interest_sheet_raw.get_highest_column()
        for col_id in range(interest_sheet_raw.get_highest_column()):
            row_data[col_id] = interest_sheet_raw.cell(row=row_id+1, column=col_id+1).value
        interest_sheet[row_id] = row_data

    # open the location sheet
    location_sheet_raw = study_book.get_sheet_by_name('Location')

    # convert cell to array
    location_sheet = [None]*location_sheet_raw.get_highest_row()
    for row_id in range(location_sheet_raw.get_highest_row()):
        row_data = [None]*location_sheet_raw.get_highest_column()
        for col_id in range(location_sheet_raw.get_highest_column()):
            row_data[col_id] = location_sheet_raw.cell(row=row_id+1, column=col_id+1).value
        location_sheet[row_id] = row_data

    # Specifying indices for the Snapshot sheet
    SNAPSHOT_DATA_ROW = 1
    SNAPSHOT_CURR_TIME_COL = 1
    TRIAL_START_TIME_ROW = 0
    TRIAL_START_TIME_COL = 0
    ACTIVATION_DATA_ROW = 2
    ACTIVATION_DATA_COL = 1
    ACTIVATION_TIME_COL = 0
    INTEREST_DATA_ROW = 1
    INTEREST_DATA_COL = 2
    LOCATION_DATA_ROW = 1
    LOCATION_DATA_COL = 2

    # record the headers
    activation_sheet_headers = activation_sheet[0][ACTIVATION_DATA_COL:]

    # 1 a. Compute the interest level for each sample point (scaled)
    sample_interest_level = interest_sheet[study_id][INTEREST_DATA_COL:]
    sample_interest_level = [level/9 for level in sample_interest_level]

    # 1 b. compute the location of the user at each sample point
    sample_loc_grid = location_sheet[study_id][LOCATION_DATA_COL:]
    sample_loc = []
    # for each samples' location
    for grid_loc in tuple(sample_loc_grid):
        # find the grid pts as indicated in the data sheet if the user is in between
        if isinstance(grid_loc, str):
            pts = grid_loc.split(",")
            # convert each to coordinate in the spatial map
            for i, pt in enumerate(pts):
                pts[i] = node_map.get_position('box_%d' % int(pt))
            # get the average
            sample_loc.append(tuple(np.mean(pts, axis=0)))
        else:
            sample_loc.append(node_map.get_position('box_%d' % grid_loc))

    #print("Participant's location: ", ['(%.2f, %.2f)' % tuple(pt[0:2]) for pt in sample_loc])

    # 2. Compute average activation among all nodes for all data points in activation sheet
    activation_win_time = []
    activation_avg_all_nodes = []
    for row_id in range(ACTIVATION_DATA_ROW, len(activation_sheet)):
        activation_win_time.append(activation_sheet[row_id][ACTIVATION_TIME_COL])
        cell_array = activation_sheet[row_id][ACTIVATION_DATA_COL:]
        # replace non-numerical values with 0
        for node_id, cell in enumerate(cell_array):
            if not isinstance(cell, (float, int)):
                cell_array[node_id] = 0.0
        activation_avg_all_nodes.append(np.mean(tuple(cell_array)))

    # 3. Acquire the activation values close to the sample points

    # acquire the start time of the trial
    trial_start_time = activation_sheet[TRIAL_START_TIME_ROW][TRIAL_START_TIME_COL]
    trial_start_time = datetime.datetime.strptime(trial_start_time, "%Y-%m-%d %H:%M:%S.%f")

    # calculate the time since the start in second
    snapshot_win_time = []
    for row_id in range(SNAPSHOT_DATA_ROW, len(snapshot_sheet)):
        snapshot_Time = datetime.datetime.strptime(snapshot_sheet[row_id][SNAPSHOT_CURR_TIME_COL], "%Y-%m-%d %H:%M:%S:%f")
        win_time = snapshot_Time - trial_start_time
        snapshot_win_time.append(win_time.total_seconds())

    # search for the windows in the activation sheet that are close to the snapshot
    sample_activation_arrays = []
    snapshot_id = 0
    for row_id in range(ACTIVATION_DATA_ROW, len(activation_sheet)):

        if activation_sheet[row_id][ACTIVATION_TIME_COL] + win_size >= snapshot_win_time[snapshot_id]:

            # cell array = current row + win_num's previous row
            sample_rows = []
            # incrementally adding rows that are at or before the snapshot
            for i in range(sample_win_num+1):
                try:
                    sample_row = activation_sheet[row_id-i+1][ACTIVATION_DATA_COL:]
                except IndexError:
                    pass
                else:
                    # making sure that there's not any non-numerical values by replacing them with 0
                    for node_id,  cell in enumerate(sample_row):
                        if not isinstance(cell, (float, int)):
                            sample_row[node_id] = 0.0
                    # append the row to the sample_rows
                    sample_rows.append(sample_row)

            # append to the for an array of arrays close to the sample points
            sample_activation_arrays.append(sample_rows)
            snapshot_id += 1

        if snapshot_id >= len(snapshot_win_time):
            break

    # 4 a. For each sample point, calculate the mean activation for each node close to the sample point
    # it is an array of average activation level for each node
    sample_avg_activation_arrays = []
    for sample_activation_array in sample_activation_arrays:
        sample_avg_activation_arrays.append(np.mean(sample_activation_array, axis=0))

    # 4 b. For each sample point, calculate the mean activation among all nodes
    sample_avg_activation_avg_all_nodes = []
    for sample_avg_activation_array in sample_avg_activation_arrays:
        sample_avg_activation_avg_all_nodes.append(np.mean(sample_avg_activation_array))

    # 4 d. For each sample point, calculate the mean average activation weighted by proximity
    sample_avg_activation_avg_proximal = []
    min_dist_to_user_arr = []
    # compute the proximity weighted avg activation at each sample point
    for sample_i, avg_arr in enumerate(sample_avg_activation_arrays):
        min_dist = [float("inf"), None]
        # sum put the weighted value of each node
        avg_sum = 0
        for node_id, node_avg in enumerate(avg_arr):
            node_name = activation_sheet_headers[node_id]
            # keep track of the closest node
            dist_to_user = node_map.get_distance(sample_loc[sample_i], node_name)
            if dist_to_user < min_dist[0]:
                min_dist[0] = dist_to_user
                min_dist[1] = node_name
            avg_sum += node_avg/dist_to_user
        sample_avg_activation_avg_proximal.append(avg_sum/len(avg_arr))
        min_dist_to_user_arr.append(min_dist)
    # print("Closest Node: ", min_dist_to_user_arr, end="\n\n")


    # 5 a. For each sample point, calculate the peak activation for each node close to the sample point
    # it is an array of average activation level for each node
    sample_peak_activation_arrays = []
    for sample_activation_array in sample_activation_arrays:
        sample_peak_activation_arrays.append(np.max(sample_activation_array, axis=0))

    # 5 b. For each sample point, calculate the mean peak activation among all nodes
    sample_peak_activation_avg_all_nodes = []
    for sample_peak_activation_array in sample_peak_activation_arrays:
        sample_peak_activation_avg_all_nodes.append(np.mean(sample_peak_activation_array))

    # 5 c. For each sample point, calculate the mean peak activation among nodes in each node type
    sample_peak_activation_avg_type = OrderedDict()
    # iterate through each cluster
    node_types = ('cbla_halfFin', 'cbla_reflex', 'cbla_light' )

    for node_type in node_types:
        # find the indices that are belonged to that node type
        node_type_indices = []
        for id, header in enumerate(activation_sheet_headers):
            if node_type in header.split('.')[1]:
                node_type_indices.append(id)

        # calculate for each sub-set of the data belonging to the cluster
        sample_peak_activation_avg_type[node_type] = []
        for sample_peak_activation_array in sample_peak_activation_arrays:
            sample_peak_activation_avg_type[node_type].append(np.mean(sample_peak_activation_array[node_type_indices]))

    # 5 d. For each sample point, calculate the mean peak activation weighted by proximity
    sample_peak_activation_avg_proximal = []
    min_dist_to_user_arr = []
    # compute the proximity weighted peak activation at each sample point
    for sample_i, peak_arr in enumerate(sample_peak_activation_arrays):
        min_dist = [float("inf"), None]
        # sum put the weighted value of each node
        peak_sum = 0
        for node_id, node_peak in enumerate(peak_arr):
            node_name = activation_sheet_headers[node_id]
            # keep track of the closest node
            dist_to_user = node_map.get_distance(sample_loc[sample_i], node_name)
            if dist_to_user < min_dist[0]:
                min_dist[0] = dist_to_user
                min_dist[1] = node_name
            peak_sum += node_peak/dist_to_user
        sample_peak_activation_avg_proximal.append(peak_sum/len(peak_arr))
        min_dist_to_user_arr.append(min_dist)
    #print("Closest Node: ", min_dist_to_user_arr, end="\n\n")

    # 6 Computer Pearson Correlations
    print("Pearson Correlation to Interest Level")
    print("========================================")

    # 6 a. Compute the Pearson correlation between interest level and average sample average activation level

    pearson_correlation = stats.pearsonr(sample_interest_level[:len(sample_avg_activation_avg_all_nodes)], sample_avg_activation_avg_all_nodes)
    correlation_stats['Average Sample Average Activation (all nodes)'].append(pearson_correlation)

    print("Average Sample Average Activation (all nodes) --- R: %f; P-Value: %f" % (pearson_correlation[0], pearson_correlation[1]))

    # 6 b. Compute the Pearson correlation between interest level and peak sample average activation level
    pearson_correlation = stats.pearsonr(sample_interest_level[:len(sample_peak_activation_avg_all_nodes)], sample_peak_activation_avg_all_nodes)
    correlation_stats['Average Sample Peak Activation (all nodes)'].append(pearson_correlation)

    print("Average Sample Peak Activation (all nodes) --- R: %f; P-Value: %f" % (pearson_correlation[0], pearson_correlation[1]))

    # 6 c. Compute the Pearson correlation between interest level and peak sample average activation level for each type
    for node_type_id, node_type_data in sample_peak_activation_avg_type.items():
        pearson_correlation = stats.pearsonr(sample_interest_level[:len(sample_peak_activation_avg_type[node_type_id])],
                                             sample_peak_activation_avg_type[node_type_id])
        correlation_stats["Average Sample Peak Activation (%s)" % node_type_id].append(pearson_correlation)

        print("Average Sample Peak Activation (%s) --- R: %f; P-Value: %f" % (node_type_id, pearson_correlation[0], pearson_correlation[1]))

    # 6 d. Compute the Pearson correlation between interest level and average sample average activation level
    pearson_correlation = stats.pearsonr(sample_interest_level[:len(sample_avg_activation_avg_proximal)], sample_avg_activation_avg_proximal)
    correlation_stats["Average Sample Average Activation (weighted by proximity)"].append(pearson_correlation)

    print("Average Sample Average Activation (weighted by proximity) --- R: %f; P-Value: %f" % (pearson_correlation[0], pearson_correlation[1]))

    # 6 e. Compute the Pearson correlation between interest level and average sample peak activation level
    pearson_correlation = stats.pearsonr(sample_interest_level[:len(sample_peak_activation_avg_proximal)], sample_peak_activation_avg_proximal)
    correlation_stats["Average Sample Peak Activation (weighted by proximity)"].append(pearson_correlation)

    print("Average Sample Peak Activation (weighted by proximity) --- R: %f; P-Value: %f" % (pearson_correlation[0], pearson_correlation[1]))

    print("\n")

    # 7. plot them

    # 7 a. plot the statistics that includes all nodes
    fig = plt.figure(1)
    plt.clf()

    # plot the sample interest level over time
    plt.plot(snapshot_win_time[:len(sample_interest_level)],
             sample_interest_level,
             "-bo", ms=10, label="Sample Interest Level")

    # plot the average activation level among all nodes
    plt.plot(activation_win_time[:len(activation_avg_all_nodes)],
             activation_avg_all_nodes,
             "-g", label="Average Activation (for all Nodes; window=%.2f)" % win_size)

    # plot the average of average sample activation level among all nodes
    plt.plot(snapshot_win_time[:len(sample_avg_activation_avg_all_nodes)],
             sample_avg_activation_avg_all_nodes,
             "-ro", ms=10, label="Average Sample Average Activation (for all Nodes; window=%.2f)" % win_size)

     # plot the average of peak sample activation level among all nodes
    plt.plot(snapshot_win_time[:len(sample_avg_activation_avg_all_nodes)],
             sample_peak_activation_avg_all_nodes,
             "-mo", ms=10, label="Average Sample Peak Activation (for all Nodes; window=%.2f)" % win_size)

    # add the x-axis label
    plt.xlabel("Time (s)")

    # add the legend
    fontP = plt_font.FontProperties()
    fontP.set_size('small')
    plt.legend( bbox_to_anchor=(1.02, 1), loc=4, prop=fontP)

    # add text about correlation
    plt.text(60, 0.1,'R=%f' % pearson_correlation[0])
    # plt.show()
    save_figure.save(fig, filename="study_%d - average_total_activation" % study_id,
                     directory="plot_figures", verbose=False)

    # 7 b. plot the statistics that includes node of each node type
    fig = plt.figure(2)
    plt.clf()
    num_plot = len(sample_peak_activation_avg_type) + 1
    line_color = iter(plt.cm.get_cmap('Set1')(np.linspace(0,1,num_plot)))

    # plot the sample interest level over time
    plt.plot(snapshot_win_time[:len(sample_interest_level)],
             sample_interest_level,
             "-o", c=next(line_color), ms=10, label="Sample Interest Level",)

    # plot the average of peak sample activation level among nodes in each node type
    for node_type_id, node_type_data in sample_peak_activation_avg_type.items():

        plt.plot(snapshot_win_time[:len(node_type_data)], node_type_data,
                 "-x", c=next(line_color), ms=5, label="Average Sample Peak Activation (for %s; window=%.2f)" % (node_type_id, win_size))

    # add the x-axis label
    plt.xlabel("Time (s)")

    # add the legend
    fontP = plt_font.FontProperties()
    fontP.set_size('small')
    plt.legend( bbox_to_anchor=(1.02, 1), loc=4, prop=fontP)

    # add text about correlation
    plt.text(60, 0.1,'R=%f' % pearson_correlation[0])
    # plt.show()
    save_figure.save(fig, filename="study_%d - average_cluster_activation" % study_id,
                     directory="plot_figures", verbose=False)

    # 7 c. plot the proximity weighted activation
    fig = plt.figure(3)
    plt.clf()
    num_plot = len(sample_peak_activation_avg_proximal) + 1
    line_color = iter(plt.cm.get_cmap('Set1')(np.linspace(0,1,num_plot)))

    # plot the sample interest level over time
    plt.plot(snapshot_win_time[:len(sample_interest_level)],
             sample_interest_level,
             "-o", c=next(line_color), ms=10, label="Sample Interest Level",)

    # plot the average of peak sample activation level weighted by proximity to the user
    plt.plot(snapshot_win_time[:len(sample_peak_activation_avg_proximal)], sample_peak_activation_avg_proximal,
             "-x", c=next(line_color), ms=5,
             label="Average Sample Peak Activation (weighted by proximity; window=%.2f)" % (win_size))

    # plot the average of peak sample activation level weighted by proximity to the user
    plt.plot(snapshot_win_time[:len(sample_avg_activation_avg_proximal)], sample_avg_activation_avg_proximal,
             "-x", c=next(line_color), ms=5,
             label="Average Sample Average Activation (weighted by proximity; window=%.2f)" % (win_size))

    # add the x-axis label
    plt.xlabel("Time (s)")

    # add the legend
    fontP = plt_font.FontProperties()
    fontP.set_size('small')
    plt.legend( bbox_to_anchor=(1.02, 1), loc=4, prop=fontP)

    # plt.show()
    save_figure.save(fig, filename="study_%d - average_proximity_weighted_activation" % study_id,
                     directory="plot_figures", verbose=False)

# Perform summary statitics on Correlation
print("Summary Statistics across Studies")
print("====================================\n")

# iterate through each correlation stats
print("Average Pearson Correlation to Interest Level")
print("-----------------------------------------------")
correlation_stats_key = sorted(tuple(correlation_stats.keys()))
for stats_key in correlation_stats_key:

    print("%s --- mean: %f; std.dev.: %f" % (stats_key, float(np.mean(correlation_stats[stats_key])),
                                             float(np.std(correlation_stats[stats_key], ddof=1))))


# return back to where the program was executed
os.chdir(program_dir)